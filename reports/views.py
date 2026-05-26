import io
from functools import wraps

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages

from scheduler.models import HorarioGenerado, SesionHorario
from academic.models import BloqueHorario, Titulacion

DIAS_LABEL = {1: 'Lunes', 2: 'Martes', 3: 'Miércoles', 4: 'Jueves', 5: 'Viernes'}
DIAS = [1, 2, 3, 4, 5]


def _es_admin(user):
    return user.is_superuser or (hasattr(user, 'rol') and user.rol in ('DECANO', 'IT', 'CONSULTOR'))


def rol_requerido(*roles):
    def decorator(view_func):
        @wraps(view_func)
        @login_required(login_url='/accounts/login/')
        def wrapped(request, *args, **kwargs):
            if request.user.is_superuser or (
                hasattr(request.user, 'rol') and request.user.rol in roles
            ):
                return view_func(request, *args, **kwargs)
            return HttpResponseForbidden(
                render(request, 'scheduler/403.html', status=403).content
            )
        return wrapped
    return decorator


@rol_requerido('DECANO', 'IT', 'CONSULTOR')
def index(request):
    """Lista de horarios disponibles para exportar."""
    titulaciones = Titulacion.objects.filter(activa=True)
    tit_filtro = request.GET.get('titulacion', '')
    sem_filtro = request.GET.get('semestre', '')

    qs = HorarioGenerado.objects.select_related('titulacion', 'curso').order_by(
        'titulacion__nombre', 'curso__numero', 'semestre'
    )
    if tit_filtro:
        qs = qs.filter(titulacion__codigo=tit_filtro)
    if sem_filtro:
        qs = qs.filter(semestre=sem_filtro)

    return render(request, 'reports/index.html', {
        'horarios':     qs,
        'titulaciones': titulaciones,
        'tit_filtro':   tit_filtro,
        'sem_filtro':   sem_filtro,
    })


@rol_requerido('DECANO', 'IT', 'CONSULTOR')
def exportar_pdf(request, pk):
    """Exporta el horario como PDF usando reportlab."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    horario = get_object_or_404(HorarioGenerado, pk=pk)
    sesiones = horario.sesiones.select_related(
        'asignatura', 'bloque', 'asignatura__asignacion__profesor__user'
    )

    bloques = list(BloqueHorario.objects.filter(activo=True).order_by('hora_inicio'))
    if getattr(horario.curso, 'es_hibrido', False):
        pass  # all blocks
    elif horario.curso.es_ultimo:
        bloques = [b for b in bloques if b.turno == 'TARDE']
    else:
        bloques = [b for b in bloques if b.turno == 'MANANA']

    grid = {dia: {b.id: None for b in bloques} for dia in DIAS}
    for s in sesiones:
        if s.dia in grid and s.bloque_id in grid[s.dia]:
            grid[s.dia][s.bloque_id] = s

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.5 * cm, leftMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Heading1'],
                                 fontSize=14, spaceAfter=4, textColor=colors.HexColor('#1a1a2e'))
    sub_style = ParagraphStyle('sub', parent=styles['Normal'],
                               fontSize=9, textColor=colors.HexColor('#666680'), spaceAfter=12)
    cell_style = ParagraphStyle('cell', parent=styles['Normal'],
                                fontSize=7.5, leading=10)
    head_style = ParagraphStyle('head', parent=styles['Normal'],
                                fontSize=8, alignment=TA_CENTER, textColor=colors.white)

    elements = []
    elements.append(Paragraph(
        f"{horario.titulacion.nombre} — {horario.curso.numero}º curso",
        title_style
    ))
    elements.append(Paragraph(
        f"Semestre {horario.semestre} · {horario.anio_academico} · {horario.get_estado_display()}",
        sub_style
    ))

    # Build table data
    header_row = [Paragraph('Bloque', head_style)] + [
        Paragraph(DIAS_LABEL[d], head_style) for d in DIAS
    ]
    rows = [header_row]

    for bloque in bloques:
        row = [
            Paragraph(
                f"<b>{bloque.nombre}</b><br/>{bloque.hora_inicio:%H:%M}–{bloque.hora_fin:%H:%M}",
                cell_style
            )
        ]
        for dia in DIAS:
            sesion = grid[dia].get(bloque.id)
            if sesion:
                txt = f"<b>{sesion.asignatura.codigo}</b><br/>{sesion.asignatura.nombre}"
                try:
                    prof = sesion.asignatura.asignacion.profesor.user
                    txt += f"<br/><font size='6'>{prof.get_full_name()}</font>"
                except Exception:
                    pass
                row.append(Paragraph(txt, cell_style))
            else:
                row.append('')
        rows.append(row)

    col_widths = [3.2 * cm] + [(landscape(A4)[0] - 1.5 * cm * 2 - 3.2 * cm) / 5] * 5

    table = Table(rows, colWidths=col_widths, repeatRows=1)
    DARK   = colors.HexColor('#1a1a2e')
    HEADER = colors.HexColor('#7c1c2e')
    LIGHT  = colors.HexColor('#f0f0f8')
    GRID   = colors.HexColor('#d0d0e0')

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HEADER),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTSIZE',   (0, 0), (-1, 0), 8),
        ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN',     (0, 0), (-1, -1), 'TOP'),
        ('BACKGROUND', (0, 1), (0, -1), LIGHT),
        ('GRID',       (0, 0), (-1, -1), 0.4, GRID),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f8fc')]),
        ('TOPPADDING',  (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 4),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
    ])
    table.setStyle(style)
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    filename = (
        f"horario_{horario.titulacion.codigo}_{horario.curso.numero}o"
        f"_S{horario.semestre}_{horario.anio_academico.replace('-', '_')}.pdf"
    )
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@rol_requerido('DECANO', 'IT', 'CONSULTOR')
def exportar_excel(request, pk):
    """Exporta el horario como Excel usando openpyxl."""
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    horario = get_object_or_404(HorarioGenerado, pk=pk)
    sesiones = horario.sesiones.select_related(
        'asignatura', 'bloque', 'asignatura__asignacion__profesor__user'
    )

    bloques = list(BloqueHorario.objects.filter(activo=True).order_by('hora_inicio'))
    if getattr(horario.curso, 'es_hibrido', False):
        pass
    elif horario.curso.es_ultimo:
        bloques = [b for b in bloques if b.turno == 'TARDE']
    else:
        bloques = [b for b in bloques if b.turno == 'MANANA']

    grid = {dia: {b.id: None for b in bloques} for dia in DIAS}
    for s in sesiones:
        if s.dia in grid and s.bloque_id in grid[s.dia]:
            grid[s.dia][s.bloque_id] = s

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{horario.titulacion.codigo} {horario.curso.numero}º S{horario.semestre}"

    # Styles
    DARK_FILL  = PatternFill('solid', fgColor='1a1a2e')
    BURG_FILL  = PatternFill('solid', fgColor='7c1c2e')
    LIGHT_FILL = PatternFill('solid', fgColor='f0f0f8')
    ALT_FILL   = PatternFill('solid', fgColor='f8f8fc')
    thin = Side(style='thin', color='d0d0e0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    white_bold = Font(bold=True, color='FFFFFF', size=9)
    dark_bold  = Font(bold=True, color='1a1a2e', size=9)
    dark_norm  = Font(color='1a1a2e', size=8)

    center = Alignment(horizontal='center', vertical='top', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='top', wrap_text=True)

    # Title row
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = (
        f"{horario.titulacion.nombre} — {horario.curso.numero}º curso · "
        f"S{horario.semestre} · {horario.anio_academico} · {horario.get_estado_display()}"
    )
    title_cell.font = Font(bold=True, color='1a1a2e', size=12)
    title_cell.fill = LIGHT_FILL
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 22

    # Header row
    headers = ['Bloque'] + [DIAS_LABEL[d] for d in DIAS]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = white_bold
        cell.fill = BURG_FILL
        cell.border = border
        cell.alignment = center
    ws.row_dimensions[2].height = 18

    # Data rows
    for r_idx, bloque in enumerate(bloques, start=3):
        fill = LIGHT_FILL if r_idx % 2 == 1 else ALT_FILL

        # Bloque label
        bloque_cell = ws.cell(
            row=r_idx, column=1,
            value=f"{bloque.nombre}\n{bloque.hora_inicio:%H:%M}–{bloque.hora_fin:%H:%M}"
        )
        bloque_cell.font = dark_bold
        bloque_cell.fill = LIGHT_FILL
        bloque_cell.border = border
        bloque_cell.alignment = left

        for d_idx, dia in enumerate(DIAS, start=2):
            sesion = grid[dia].get(bloque.id)
            cell = ws.cell(row=r_idx, column=d_idx)
            cell.border = border
            cell.alignment = left

            if sesion:
                prof_name = ''
                try:
                    prof = sesion.asignatura.asignacion.profesor.user
                    prof_name = f"\n{prof.get_full_name()}"
                except Exception:
                    pass
                cell.value = (
                    f"{sesion.asignatura.codigo}\n"
                    f"{sesion.asignatura.nombre}"
                    f"{prof_name}"
                )
                cell.font = dark_norm
                cell.fill = fill
            else:
                cell.fill = fill

        ws.row_dimensions[r_idx].height = 40

    # Column widths
    ws.column_dimensions['A'].width = 16
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 24

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = (
        f"horario_{horario.titulacion.codigo}_{horario.curso.numero}o"
        f"_S{horario.semestre}_{horario.anio_academico.replace('-', '_')}.xlsx"
    )
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
